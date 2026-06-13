#!/usr/bin/env python3
"""
AquaVir-KB 核心表 → 多 Sheet Excel 导出
输出: public_downloads/AquaVir-KB_core_tables.xlsx
"""

import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = os.path.join(os.path.dirname(__file__), "crustacean_virus_core.db")
OUT_DIR = os.path.join(os.path.dirname(__file__), "public_downloads")
OUT_PATH = os.path.join(OUT_DIR, "AquaVir-KB_core_tables_v2.xlsx")

# 7 张核心表的配置: (Sheet名, 数据源SQL, 表说明)
SHEETS = [
    ("病毒物种", "SELECT * FROM public_virus_master", "病毒物种主表"),
    ("病毒分离株", "SELECT * FROM public_viral_isolates", "病毒分离株 / GenBank 记录"),
    ("病毒蛋白", "SELECT * FROM viral_proteins", "病毒蛋白质序列"),
    ("宿主信息", "SELECT * FROM public_crustacean_hosts", "水生无脊椎动物宿主"),
    ("证据记录", "SELECT * FROM public_evidence_records", "结构化文献证据"),
    ("感染记录", "SELECT * FROM infection_records", "病毒感染记录"),
    ("文献", "SELECT * FROM public_ref_literatures", "参考文献"),
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
CELL_FONT = Font(name="微软雅黑", size=9)
CELL_ALIGNMENT = Alignment(vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# Excel XML 不允许的字符 (包括控制字符 0x00-0x08, 0x0B-0x0C, 0x0E-0x1F)
ILLEGAL_CHARS = set(
    list(range(0x00, 0x09)) + [0x0B, 0x0C] + list(range(0x0E, 0x20))
)

def clean_value(val):
    """清除 Excel 不接受的字符"""
    if isinstance(val, str):
        # 过滤掉非法字符，用空格替换
        cleaned = []
        for ch in val:
            if ord(ch) in ILLEGAL_CHARS:
                cleaned.append(" ")
            else:
                cleaned.append(ch)
        return "".join(cleaned)
    return val


def auto_width(ws, min_width=6, max_width=50):
    """根据内容自适应列宽"""
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells[:100]:  # 取前 100 行采样
            if cell.value:
                # 中文字符算 2 个宽度
                val = str(cell.value)
                char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, char_len)
        width = max(min(max_len + 2, max_width), min_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def export():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout = 30000")

    wb = Workbook()
    # 删除默认空 sheet
    wb.remove(wb.active)

    for sheet_name, sql, description in SHEETS:
        print(f"Exporting: {sheet_name} ...", end=" ", flush=True)

        # 获取列名
        cursor = conn.execute(f"{sql} LIMIT 0")
        columns = [d[0] for d in cursor.description]

        # 读取数据
        rows = conn.execute(sql).fetchall()

        ws = wb.create_sheet(title=sheet_name)

        # 写表头行
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=clean_value(col_name))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        # 批量写数据行
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=clean_value(value))
                cell.font = CELL_FONT
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER

        # 格式: 冻结首行 + 筛选 + 自适应列宽
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows)+1}"
        auto_width(ws)

        print(f"{len(rows):,} rows, {len(columns)} cols [OK]")

    conn.close()

    # 写第一个 sheet 放说明
    overview = wb.create_sheet(title="说明", index=0)
    overview_data = [
        ["AquaVir-KB 核心数据表", ""],
        ["", ""],
        ["Sheet 名", "内容说明"],
    ]
    for sheet_name, _, description in SHEETS:
        overview_data.append([sheet_name, description])

    for row_idx, (col_a, col_b) in enumerate(overview_data, 1):
        a = overview.cell(row=row_idx, column=1, value=clean_value(col_a))
        b = overview.cell(row=row_idx, column=2, value=clean_value(col_b))
        if row_idx == 1:
            a.font = Font(name="微软雅黑", size=14, bold=True)
            b.font = Font(name="微软雅黑", size=14, bold=True)
        elif row_idx == 3:
            a.font = Font(name="微软雅黑", size=10, bold=True)
            b.font = Font(name="微软雅黑", size=10, bold=True)
        else:
            a.font = Font(name="微软雅黑", size=10)
            b.font = Font(name="微软雅黑", size=10)
    overview.column_dimensions["A"].width = 20
    overview.column_dimensions["B"].width = 45

    # 确保输出目录存在
    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(OUT_PATH)

    size_mb = os.path.getsize(OUT_PATH) / (1024 * 1024)
    print(f"\nDone: {OUT_PATH}")
    print(f"Size: {size_mb:.1f} MB")


if __name__ == "__main__":
    export()
