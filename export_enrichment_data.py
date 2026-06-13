"""
Export all enrichment data from SQLite to downloadable files.

Generates:
  downloads/enrichment/
    all_enrichment_data.xlsx       (multi-sheet Excel)
    kegg_annotations.csv
    kegg_pathways.csv
    kegg_protein_pathways.csv
    viralzone_families.csv
    interpro_annotations.csv
    geo_datasets.csv
    gbif_occurrences.csv
    gbif_species_summary.csv
    epmc_literature.csv
    pride_datasets.csv
    alphafold_structures.csv
    string_interactions.csv
    biorxiv_preprints.csv
    obis_occurrences.csv
    enrichment_summary.json        (stats overview)
    enrichment_summary.csv         (stats overview)

Usage:
    python export_enrichment_data.py              # export all
    python export_enrichment_data.py --format csv # CSV only
    python export_enrichment_data.py --format xlsx # Excel only
"""
import csv
import json
import sqlite3
from pathlib import Path
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "crustacean_virus_core.db"
OUTPUT_DIR = BASE_DIR / "downloads" / "enrichment"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TABLES = {
    "kegg_annotations": {
        "name": "KEGG 通路注释",
        "description": "KEGG Orthology annotations with EC numbers, KO identifiers, and pathway mappings",
    },
    "kegg_pathways": {
        "name": "KEGG 通路列表",
        "description": "KEGG pathway maps linked to viral proteins",
    },
    "kegg_protein_pathways": {
        "name": "KEGG 蛋白-通路链接",
        "description": "Protein-to-pathway links via KEGG Orthology",
    },
    "viralzone_families": {
        "name": "ViralZone 病毒科",
        "description": "Virus family factsheets from ViralZone (SIB)",
    },
    "interpro_annotations": {
        "name": "InterPro 蛋白结构域",
        "description": "Protein domain annotations from InterPro via EBI",
    },
    "geo_datasets": {
        "name": "GEO 转录组数据集",
        "description": "NCBI GEO transcriptomics datasets for crustacean virus studies",
    },
    "gbif_occurrences": {
        "name": "GBIF 宿主出现记录",
        "description": "Host species occurrence records from GBIF for distribution mapping",
    },
    "gbif_species_summary": {
        "name": "GBIF 物种分布摘要",
        "description": "Aggregated species distribution summaries from GBIF",
    },
    "epmc_literature": {
        "name": "Europe PMC 文献",
        "description": "Enriched literature metadata from Europe PMC",
    },
    "pride_datasets": {
        "name": "PRIDE 蛋白质组数据集",
        "description": "Mass spectrometry proteomics datasets from PRIDE/ProteomeXchange",
    },
    "uniprot_structures": {
        "name": "AlphaFold 蛋白结构",
        "description": "AlphaFold DB and PDB protein 3D structure predictions and experimental data",
    },
    "string_interactions": {
        "name": "STRING 蛋白互作",
        "description": "Host immune protein-protein interaction networks from STRING",
    },
    "biorxiv_preprints": {
        "name": "bioRxiv 预印本",
        "description": "Preprint metadata from bioRxiv and medRxiv via Crossref",
    },
    "obis_occurrences": {
        "name": "OBIS 海洋宿主记录",
        "description": "Marine host species occurrence records from OBIS",
    },
    "host_ecological_traits": {
        "name": "宿主生态特征",
        "description": "Ecological trait data from EOL/FishBase",
    },
    "host_biology_profiles": {
        "name": "宿主生物学档案",
        "description": "Host species biology profiles (habitat, temperature, fecundity, etc.)",
    },
}

# Columns to exclude from export (internal/raw JSON blobs)
EXCLUDE_COLS = {"raw_json", "raw_sections"}


def export_table_to_csv(conn: sqlite3.Connection, table: str, path: Path) -> int:
    """Export a single table to CSV. Returns row count."""
    cursor = conn.execute(f"SELECT * FROM {table}")
    rows = cursor.fetchall()
    if not rows:
        path.write_text("", encoding="utf-8")
        return 0

    col_names = [desc[0] for desc in cursor.description if desc[0] not in EXCLUDE_COLS]
    col_indices = [i for i, desc in enumerate(cursor.description) if desc[0] not in EXCLUDE_COLS]

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(col_names)
        for row in rows:
            writer.writerow([str(row[i]) if row[i] is not None else "" for i in col_indices])

    return len(rows)


def export_to_xlsx(conn: sqlite3.Connection, stats: dict) -> Path:
    """Export all tables to a multi-sheet Excel file. Requires openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[warn] openpyxl not installed. Run: pip install openpyxl")
        print("[warn] Skipping Excel export. CSV files are available instead.")
        return None

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Summary sheet
    ws_summary = wb.create_sheet("富集数据总览")
    ws_summary.append(["数据表", "记录数", "说明"])
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 60

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0f766e", end_color="0f766e", fill_type="solid")

    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for table, info in TABLES.items():
        cnt = stats.get(table, 0)
        ws_summary.append([info["name"], cnt, info["description"]])

    # Data sheets
    for table, info in TABLES.items():
        try:
            cursor = conn.execute(f"SELECT * FROM {table}")
            rows = cursor.fetchall()
        except Exception:
            continue

        if not rows:
            continue

        col_names = [desc[0] for desc in cursor.description if desc[0] not in EXCLUDE_COLS]
        col_indices = [i for i, desc in enumerate(cursor.description) if desc[0] not in EXCLUDE_COLS]

        # Truncate sheet name to 31 chars (Excel limit)
        sheet_name = info["name"][:31]
        ws = wb.create_sheet(sheet_name)

        # Header row
        ws.append(col_names)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color="134e4a", end_color="134e4a", fill_type="solid")

        # Data rows (limit to 10k rows per sheet for performance)
        for row in rows[:10000]:
            ws.append([str(row[i]) if row[i] is not None else "" for i in col_indices])

        # Auto-width columns
        for col_idx, name in enumerate(col_names, 1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max(len(str(name)) * 1.2, 10), 40)

    xlsx_path = OUTPUT_DIR / "all_enrichment_data.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


def export_summary(stats: dict) -> dict:
    """Generate summary JSON and CSV."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary = {
        "export_time": timestamp,
        "database": str(DB_PATH),
        "tables": {}
    }

    for table, info in TABLES.items():
        cnt = stats.get(table, 0)
        summary["tables"][table] = {
            "name": info["name"],
            "description": info["description"],
            "record_count": cnt,
        }

    # JSON
    json_path = OUTPUT_DIR / "enrichment_summary.json"
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    # CSV summary
    csv_path = OUTPUT_DIR / "enrichment_summary.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["表名", "名称", "记录数", "说明"])
        for table, info in TABLES.items():
            writer.writerow([table, info["name"], stats.get(table, 0), info["description"]])

    return summary


def main():
    print("=" * 60)
    print("Crustacean Virus DB - Enrichment Data Export")
    print(f"Output: {OUTPUT_DIR}")
    print("=" * 60)

    import argparse
    parser = argparse.ArgumentParser(description="Export enrichment data to local files")
    parser.add_argument("--format", choices=["csv", "xlsx", "all"], default="all",
                        help="Export format (default: all)")
    args = parser.parse_args()

    conn = sqlite3.connect(str(DB_PATH))
    stats = {}

    # Export CSV
    if args.format in ("csv", "all"):
        print("\n--- CSV Export ---")
        total = 0
        for table, info in TABLES.items():
            try:
                path = OUTPUT_DIR / f"{table}.csv"
                cnt = export_table_to_csv(conn, table, path)
                stats[table] = cnt
                total += cnt
                marker = "[OK]" if cnt > 0 else "[EMPTY]"
                print(f"  {marker} {info['name']:30s} → {path.name:40s} ({cnt:,} rows)")
            except Exception as e:
                print(f"  [SKIP] {info['name']:30s} — {e}")
                stats[table] = 0
        print(f"  Total CSV rows exported: {total:,}")

    # Export Excel
    if args.format in ("xlsx", "all"):
        print("\n--- Excel Export ---")
        xlsx_path = export_to_xlsx(conn, stats)
        if xlsx_path:
            print(f"  [OK] Multi-sheet Excel: {xlsx_path}")
            print(f"  Size: {xlsx_path.stat().st_size / 1024:.0f} KB")
        else:
            print("  [SKIP] Excel requires openpyxl: pip install openpyxl")

    # Summary
    print("\n--- Summary ---")
    summary = export_summary(stats)
    print(f"  [OK] {OUTPUT_DIR / 'enrichment_summary.json'}")
    print(f"  [OK] {OUTPUT_DIR / 'enrichment_summary.csv'}")
    print(f"\n  Total tables exported: {len([t for t, c in stats.items() if c > 0])}")
    print(f"  Total records exported: {sum(stats.values()):,}")

    conn.close()
    print(f"\n[done] All files saved to: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
